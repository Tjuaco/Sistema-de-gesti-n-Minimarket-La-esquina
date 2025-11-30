from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.core.cache import cache
from usuarios.permissions import PuedeVentas
from .models import Venta, DetalleVenta
from .serializers import VentaSerializer, CrearVentaSerializer


class VentaViewSet(viewsets.ModelViewSet):
    queryset = Venta.objects.prefetch_related('items__producto').all()
    serializer_class = VentaSerializer
    search_fields = ['numero_boleta']
    permission_classes = [IsAuthenticated, PuedeVentas]  # Cajero y Administrador

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return CrearVentaSerializer
        return VentaSerializer

    def get_queryset(self):
        queryset = Venta.objects.prefetch_related('items__producto').all()
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)

        if fecha_desde:
            queryset = queryset.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha__lte=fecha_hasta)

        return queryset.order_by('-fecha')

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        venta = serializer.save()
        
        # Invalidar caché de productos activos después de una venta
        cache.delete('productos_activos')
        
        return Response(
            VentaSerializer(venta).data,
            status=status.HTTP_201_CREATED
        )

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=False)
        serializer.is_valid(raise_exception=True)
        venta = serializer.save()
        
        return Response(
            VentaSerializer(venta).data,
            status=status.HTTP_200_OK
        )

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        """Eliminar una venta y revertir los cambios de stock"""
        try:
            instance = self.get_object()
            usuario = request.user.username if request.user.is_authenticated else 'Cajero'
            
            # Revertir los cambios de stock de cada item
            from inventario.models import MovimientoStock
            
            for detalle in instance.items.all():
                producto = detalle.producto
                stock_actual = producto.stock_actual
                stock_nuevo = stock_actual + detalle.cantidad  # Revertir: sumar lo que se había restado
                
                # Revertir stock
                producto.stock_actual = stock_nuevo
                producto.save()
                
                # Registrar movimiento de reversión
                MovimientoStock.objects.create(
                    producto=producto,
                    tipo='AJUSTE',
                    cantidad=detalle.cantidad,  # Positivo porque estamos revirtiendo
                    stock_anterior=stock_actual,
                    stock_nuevo=stock_nuevo,
                    motivo=f'Eliminación de Venta #{instance.id}',
                    usuario=usuario
                )
            
            # Eliminar la venta (los detalles se eliminan en cascada)
            self.perform_destroy(instance)
            
            return Response(
                {'mensaje': 'Venta eliminada exitosamente. Los cambios de stock han sido revertidos.'},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {'error': f'Error al eliminar la venta: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def exportar_csv(self, request):
        """Exportar ventas a Excel con diseño mejorado"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from datetime import datetime
        
        queryset = self.get_queryset()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # Estilos mejorados (mismo estilo que movimientos)
        title_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        data_font = Font(size=10)
        number_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Título del reporte
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'REPORTE DE VENTAS'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.border = border_style
        ws.row_dimensions[1].height = 25
        
        # Información del reporte
        ws.merge_cells('A2:H2')
        info_cell = ws['A2']
        info_cell.value = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de ventas: {queryset.count()}'
        info_cell.font = Font(size=9, italic=True, color='666666')
        info_cell.alignment = center_align
        ws.row_dimensions[2].height = 18
        
        # Encabezados
        headers = ['ID', 'Número Boleta', 'Fecha', 'Registrado por', 'Items', 'Total', 'Usuario', 'Observaciones']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        ws.row_dimensions[3].height = 20
        
        # Datos
        for row_num, venta in enumerate(queryset, 4):
            # ID
            cell = ws.cell(row=row_num, column=1, value=venta.id)
            cell.border = border_style
            cell.font = data_font
            cell.alignment = center_align
            
            # Número Boleta
            cell = ws.cell(row=row_num, column=2, value=venta.numero_boleta or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Fecha
            cell = ws.cell(row=row_num, column=3, value=venta.fecha.strftime('%d/%m/%Y %H:%M:%S'))
            cell.border = border_style
            cell.font = data_font
            cell.alignment = center_align
            
            # Registrado por (usuario)
            cell = ws.cell(row=row_num, column=4, value=venta.usuario or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Items (cantidad de items)
            cantidad_items = venta.items.count()
            cell = ws.cell(row=row_num, column=5, value=cantidad_items)
            cell.border = border_style
            cell.font = number_font
            cell.alignment = center_align
            
            # Total
            cell = ws.cell(row=row_num, column=6, value=float(venta.total))
            cell.border = border_style
            cell.font = Font(size=10, bold=True)
            cell.alignment = right_align
            cell.number_format = '#,##0'
            
            # Usuario (duplicado para compatibilidad)
            cell = ws.cell(row=row_num, column=7, value=venta.usuario or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Observaciones
            cell = ws.cell(row=row_num, column=8, value=venta.observaciones or '')
            cell.border = border_style
            cell.font = Font(size=9)
            cell.alignment = left_align
            
            ws.row_dimensions[row_num].height = 18
        
        # Ajustar ancho de columnas
        column_widths = [8, 18, 18, 18, 10, 15, 18, 35]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


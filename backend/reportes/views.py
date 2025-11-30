from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Sum, Count, F, Q, Avg, Min, Value
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import datetime, timedelta
import csv
from decimal import Decimal
from usuarios.permissions import PuedeReportes

from ventas.models import Venta, DetalleVenta
from compras.models import Compra, DetalleCompra
from inventario.models import Producto, Proveedor

# Importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportesViewSet(viewsets.ViewSet):
    """ViewSet para generar reportes"""
    permission_classes = [IsAuthenticated, PuedeReportes]  # Solo Administrador

    @action(detail=False, methods=['get'])
    def ventas_diarias(self, request):
        """Reporte de ventas diarias"""
        fecha = request.query_params.get('fecha', None)
        try:
            if fecha:
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
            else:
                fecha_obj = timezone.now().date()
            
            fecha_desde = timezone.make_aware(datetime.combine(fecha_obj, datetime.min.time()))
            fecha_hasta = timezone.make_aware(datetime.combine(fecha_obj, datetime.max.time()))
        except ValueError:
            return Response(
                {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )

        ventas = Venta.objects.filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta)
        
        total_ventas = ventas.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_ventas = ventas.count()
        
        # Detalle por producto
        detalles = DetalleVenta.objects.filter(
            venta__fecha__gte=fecha_desde,
            venta__fecha__lte=fecha_hasta
        ).values('producto__nombre', 'producto__codigo').annotate(
            cantidad_vendida=Sum('cantidad'),
            total_vendido=Sum(F('cantidad') * F('precio_unitario')),
            margen_ganancia=Sum((F('cantidad') * F('precio_unitario')) - (F('producto__costo') * F('cantidad')))
        ).order_by('-total_vendido')

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv' or formato == 'excel':
            return self._generar_excel_ventas_diarias(ventas, detalles, fecha_obj)

        return Response({
            'fecha': fecha_obj.isoformat(),
            'total_ventas': float(total_ventas),
            'cantidad_ventas': cantidad_ventas,
            'detalle_productos': list(detalles)
        })

    @action(detail=False, methods=['get'])
    def compras_mensuales(self, request):
        """Reporte de compras mensuales"""
        try:
            año = int(request.query_params.get('año', timezone.now().year))
            mes = int(request.query_params.get('mes', timezone.now().month))
            
            # Validar rango de mes
            if mes < 1 or mes > 12:
                return Response(
                    {'error': 'Mes debe estar entre 1 y 12'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar rango de año
            if año < 2000 or año > 2100:
                return Response(
                    {'error': 'Año debe estar entre 2000 y 2100'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except ValueError:
            return Response(
                {'error': 'Año y mes deben ser números válidos'},
                status=status.HTTP_400_BAD_REQUEST
            )

        fecha_inicio = datetime(año, mes, 1).date()
        if mes == 12:
            fecha_fin = datetime(año + 1, 1, 1).date() - timedelta(days=1)
        else:
            fecha_fin = datetime(año, mes + 1, 1).date() - timedelta(days=1)

        fecha_desde = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        fecha_hasta = timezone.make_aware(datetime.combine(fecha_fin, datetime.max.time()))

        compras = Compra.objects.filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta)
        
        total_compras = compras.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_compras = compras.count()
        
        # Top productos comprados
        top_productos = DetalleCompra.objects.filter(
            compra__fecha__gte=fecha_desde,
            compra__fecha__lte=fecha_hasta
        ).values('producto__nombre', 'producto__codigo').annotate(
            cantidad_comprada=Sum('cantidad'),
            total_comprado=Sum(F('cantidad') * F('costo_unitario'))
        ).order_by('-total_comprado')[:10]

        # Compras por proveedor
        compras_por_proveedor = compras.values('proveedor__nombre', 'proveedor__id').annotate(
            total_comprado=Sum('total'),
            cantidad_compras=Count('id')
        ).order_by('-total_comprado')

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv':
            return self._generar_csv_compras_mensuales(compras, top_productos, compras_por_proveedor, año, mes)
        elif formato == 'excel':
            return self._generar_excel_compras_mensuales(compras, top_productos, compras_por_proveedor, año, mes)

        return Response({
            'año': año,
            'mes': mes,
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': fecha_fin.isoformat(),
            'total_compras': float(total_compras),
            'cantidad_compras': cantidad_compras,
            'top_productos': list(top_productos),
            'compras_por_proveedor': list(compras_por_proveedor)
        })

    @action(detail=False, methods=['get'])
    def reporte_proveedores(self, request):
        """Reporte de proveedores con estadísticas"""
        from inventario.models import Proveedor
        
        proveedores = Proveedor.objects.filter(activo=True).annotate(
            cantidad_productos=Count('producto', distinct=True),
            total_compras=Sum(
                'compra__total',
                filter=Q(compra__fecha__gte=timezone.now() - timedelta(days=30))
            ),
            cantidad_compras=Count(
                'compra',
                filter=Q(compra__fecha__gte=timezone.now() - timedelta(days=30))
            )
        ).order_by('-total_compras')

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv':
            return self._generar_csv_proveedores(proveedores)
        elif formato == 'excel':
            return self._generar_excel_proveedores(proveedores)

        return Response({
            'proveedores': [
                {
                    'id': p.id,
                    'nombre': p.nombre,
                    'rut': p.rut or '',
                    'contacto': p.contacto or '',
                    'telefono': p.telefono or '',
                    'email': p.email or '',
                    'cantidad_productos': p.cantidad_productos or 0,
                    'total_compras_30dias': float(p.total_compras or 0),
                    'cantidad_compras_30dias': p.cantidad_compras or 0,
                }
                for p in proveedores
            ]
        })

    @action(detail=False, methods=['get'])
    def reporte_productos(self, request):
        """Reporte de productos con estadísticas"""
        productos = Producto.objects.filter(activo=True).select_related('categoria').annotate(
            cantidad_vendida_30dias=Sum(
                'detalleventa__cantidad',
                filter=Q(detalleventa__venta__fecha__gte=timezone.now() - timedelta(days=30))
            ),
            cantidad_comprada_30dias=Sum(
                'detallecompra__cantidad',
                filter=Q(detallecompra__compra__fecha__gte=timezone.now() - timedelta(days=30))
            )
        )

        # Productos con stock bajo - QuerySet para Excel/CSV
        productos_bajo_stock_qs = productos.filter(
            stock_actual__lte=F('stock_minimo')
        )
        
        # Productos más vendidos (30 días) - QuerySet para Excel/CSV
        productos_mas_vendidos_qs = productos.filter(
            cantidad_vendida_30dias__gt=0
        ).order_by('-cantidad_vendida_30dias')[:10]

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv':
            return self._generar_csv_productos(productos, productos_bajo_stock_qs, productos_mas_vendidos_qs)
        elif formato == 'excel':
            return self._generar_excel_productos(productos, productos_bajo_stock_qs, productos_mas_vendidos_qs)

        # Para JSON, convertir a lista de diccionarios
        productos_bajo_stock_list = list(productos_bajo_stock_qs.values(
            'codigo', 'nombre', 'stock_actual', 'stock_minimo', 'categoria__nombre'
        ))
        
        productos_mas_vendidos_list = list(productos_mas_vendidos_qs.values(
            'codigo', 'nombre', 'cantidad_vendida_30dias', 'precio_venta'
        ))

        return Response({
            'productos_bajo_stock': productos_bajo_stock_list,
            'productos_mas_vendidos': productos_mas_vendidos_list,
            'total_productos_activos': productos.count(),
            'total_productos_bajo_stock': len(productos_bajo_stock_list),
        })

    @action(detail=False, methods=['get'])
    def ventas_semanales(self, request):
        """Reporte de ventas semanales"""
        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        try:
            if fecha_inicio_str:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            else:
                # Lunes de esta semana
                today = timezone.now().date()
                fecha_inicio = today - timedelta(days=today.weekday())
        except ValueError:
            return Response(
                {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )

        fecha_fin = fecha_inicio + timedelta(days=6)
        fecha_desde = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        fecha_hasta = timezone.make_aware(datetime.combine(fecha_fin, datetime.max.time()))

        ventas = Venta.objects.filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta)
        
        total_ventas = ventas.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_ventas = ventas.count()
        
        # Ventas por día
        ventas_por_dia = ventas.annotate(
            dia=TruncDate('fecha')
        ).values('dia').annotate(
            total_dia=Sum('total'),
            cantidad_dia=Count('id')
        ).order_by('dia')

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv':
            return self._generar_csv_ventas_semanales(ventas, ventas_por_dia, fecha_inicio, fecha_fin)
        elif formato == 'excel':
            return self._generar_excel_ventas_semanales(ventas, ventas_por_dia, fecha_inicio, fecha_fin)

        return Response({
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': fecha_fin.isoformat(),
            'total_ventas': float(total_ventas),
            'cantidad_ventas': cantidad_ventas,
            'ventas_por_dia': list(ventas_por_dia)
        })

    @action(detail=False, methods=['get'])
    def ventas_mensuales(self, request):
        """Reporte de ventas mensuales"""
        try:
            año = int(request.query_params.get('año', timezone.now().year))
            mes = int(request.query_params.get('mes', timezone.now().month))
            
            # Validar rango de mes
            if mes < 1 or mes > 12:
                return Response(
                    {'error': 'Mes debe estar entre 1 y 12'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar rango de año
            if año < 2000 or año > 2100:
                return Response(
                    {'error': 'Año debe estar entre 2000 y 2100'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except ValueError:
            return Response(
                {'error': 'Año y mes deben ser números válidos'},
                status=status.HTTP_400_BAD_REQUEST
            )

        fecha_inicio = datetime(año, mes, 1).date()
        if mes == 12:
            fecha_fin = datetime(año + 1, 1, 1).date() - timedelta(days=1)
        else:
            fecha_fin = datetime(año, mes + 1, 1).date() - timedelta(days=1)

        fecha_desde = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        fecha_hasta = timezone.make_aware(datetime.combine(fecha_fin, datetime.max.time()))

        ventas = Venta.objects.filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta)
        
        total_ventas = ventas.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_ventas = ventas.count()

        # Top productos
        top_productos = DetalleVenta.objects.filter(
            venta__fecha__gte=fecha_desde,
            venta__fecha__lte=fecha_hasta
        ).values('producto__nombre', 'producto__codigo').annotate(
            cantidad_vendida=Sum('cantidad'),
            total_vendido=Sum(F('cantidad') * F('precio_unitario'))
        ).order_by('-total_vendido')[:10]

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv':
            return self._generar_csv_ventas_mensuales(ventas, top_productos, año, mes)
        elif formato == 'excel':
            return self._generar_excel_ventas_mensuales(ventas, top_productos, año, mes)

        return Response({
            'año': año,
            'mes': mes,
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': fecha_fin.isoformat(),
            'total_ventas': float(total_ventas),
            'cantidad_ventas': cantidad_ventas,
            'top_productos': list(top_productos)
        })

    @action(detail=False, methods=['get'])
    def margen_productos(self, request):
        """Reporte de margen por producto"""
        productos = Producto.objects.filter(activo=True).annotate(
            margen_calculado=((F('precio_venta') - F('costo')) / F('costo')) * 100,
            ganancia_unitaria=F('precio_venta') - F('costo')
        ).order_by('-margen_calculado')

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv':
            return self._generar_csv_margen_productos(productos)
        elif formato == 'excel':
            return self._generar_excel_margen_productos(productos)

        return Response({
            'productos': [
                {
                    'codigo': p.codigo,
                    'nombre': p.nombre,
                    'costo': float(p.costo),
                    'precio_venta': float(p.precio_venta),
                    'margen_porcentaje': float(p.margen_calculado),
                    'ganancia_unitaria': float(p.ganancia_unitaria),
                    'stock_actual': p.stock_actual
                }
                for p in productos
            ]
        })

    @action(detail=False, methods=['get'])
    def rotacion_inventario(self, request):
        """Reporte de rotación de inventario"""
        productos = Producto.objects.filter(activo=True).annotate(
            cantidad_vendida=Sum(
                'detalleventa__cantidad',
                filter=Q(detalleventa__venta__fecha__gte=timezone.now() - timedelta(days=30))
            )
        )

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv':
            return self._generar_csv_rotacion(productos)
        elif formato == 'excel':
            return self._generar_excel_rotacion(productos)

        return Response({
            'productos': [
                {
                    'codigo': p.codigo,
                    'nombre': p.nombre,
                    'stock_actual': p.stock_actual,
                    'cantidad_vendida_30dias': p.cantidad_vendida or 0,
                    'rotacion': (p.cantidad_vendida or 0) / p.stock_actual if p.stock_actual > 0 else 0
                }
                for p in productos if p.stock_actual > 0
            ]
        })

    @action(detail=False, methods=['get'])
    def quiebres_semana(self, request):
        """Reporte de quiebres de stock por semana"""
        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        else:
            # Lunes de esta semana
            today = timezone.now().date()
            fecha_inicio = today - timedelta(days=today.weekday())

        fecha_fin = fecha_inicio + timedelta(days=6)
        fecha_desde = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        fecha_hasta = timezone.make_aware(datetime.combine(fecha_fin, datetime.max.time()))

        # Buscar productos que tuvieron stock 0 o negativo durante la semana
        from inventario.models import MovimientoStock
        
        quiebres = MovimientoStock.objects.filter(
            fecha__gte=fecha_desde,
            fecha__lte=fecha_hasta,
            stock_nuevo__lte=0
        ).values('producto__codigo', 'producto__nombre').annotate(
            cantidad_quiebres=Count('id'),
            fecha_primer_quiebre=Min('fecha'),
            stock_minimo_alcanzado=Min('stock_nuevo')
        ).order_by('-cantidad_quiebres')

        formato = request.query_params.get('formato', 'json')
        if formato == 'csv' or formato == 'excel':
            return self._generar_excel_quiebres(quiebres, fecha_inicio, fecha_fin)

        return Response({
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': fecha_fin.isoformat(),
            'total_quiebres': quiebres.count(),
            'quiebres': list(quiebres)
        })

    def _generar_csv_ventas_diarias(self, ventas, detalles, fecha):
        """Método legacy - ahora usa Excel"""
        return self._generar_excel_ventas_diarias(ventas, detalles, fecha)
    
    def _generar_excel_ventas_diarias(self, ventas, detalles, fecha):
        """Generar Excel de ventas diarias con diseño profesional"""
        if not OPENPYXL_AVAILABLE:
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="ventas_diarias_{fecha}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Reporte de Ventas Diarias'])
            writer.writerow(['Fecha', fecha.isoformat()])
            writer.writerow(['Total Ventas', f"${ventas.aggregate(total=Sum('total'))['total'] or Decimal('0.00'):,.0f}"])
            writer.writerow(['Cantidad de Ventas', ventas.count()])
            writer.writerow([])
            writer.writerow(['Código Producto', 'Nombre Producto', 'Cantidad Vendida', 'Total Vendido', 'Margen Ganancia'])
            for detalle in detalles:
                writer.writerow([
                    detalle.get('producto__codigo', ''),
                    detalle.get('producto__nombre', ''),
                    detalle.get('cantidad_vendida', 0),
                    f"${float(detalle.get('total_vendido', 0)):,.0f}",
                    f"${float(detalle.get('margen_ganancia', 0) or 0):,.0f}"
                ])
            return response
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas Diarias"
        
        total_ventas = ventas.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_ventas = ventas.count()
        
        estilos = self._obtener_estilos_excel('1976D2')
        if not estilos:
            return self._generar_csv_ventas_diarias(ventas, detalles, fecha)
        
        row_start = self._aplicar_titulo_excel(
            ws, 
            f'REPORTE DE VENTAS DIARIAS - {fecha.strftime("%d/%m/%Y")}',
            f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Total: ${float(total_ventas):,.0f} | Cantidad de ventas: {cantidad_ventas}',
            5,
            estilos
        )
        
        headers = ['Código', 'Nombre Producto', 'Cantidad Vendida', 'Total Vendido', 'Margen Ganancia']
        self._aplicar_headers_excel(ws, headers, row_start, estilos)
        
        for idx, detalle in enumerate(detalles, row_start + 1):
            # Código
            cell = ws.cell(row=idx, column=1, value=detalle.get('producto__codigo', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Nombre
            cell = ws.cell(row=idx, column=2, value=detalle.get('producto__nombre', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Cantidad
            cell = ws.cell(row=idx, column=3, value=detalle.get('cantidad_vendida', 0))
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['right_align']
            
            # Total
            total = float(detalle.get('total_vendido', 0))
            cell = ws.cell(row=idx, column=4, value=f"${total:,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            # Margen
            margen = float(detalle.get('margen_ganancia', 0) or 0)
            cell = ws.cell(row=idx, column=5, value=f"${margen:,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            if margen > 0:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            
            ws.row_dimensions[idx].height = 18
        
        column_widths = [15, 35, 18, 18, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'ventas_diarias_{fecha.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def _generar_csv_ventas_semanales(self, ventas, ventas_por_dia, fecha_inicio, fecha_fin):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="ventas_semanales_{fecha_inicio}.csv"'
        
        total_semana = ventas.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_semana = ventas.count()
        
        writer = csv.writer(response)
        writer.writerow(['Reporte de Ventas Semanales'])
        writer.writerow(['Período', f'{fecha_inicio} a {fecha_fin}'])
        writer.writerow(['Total Semanal', f"${float(total_semana):,.0f}"])
        writer.writerow(['Cantidad de Ventas', cantidad_semana])
        writer.writerow([])
        writer.writerow(['Fecha', 'Total Ventas', 'Cantidad Ventas'])
        
        for dia in ventas_por_dia:
            writer.writerow([
                str(dia.get('dia', '')),
                f"${float(dia.get('total_dia', 0)):,.0f}",
                dia.get('cantidad_dia', 0)
            ])
        
        return response

    def _generar_csv_ventas_mensuales(self, ventas, top_productos, año, mes):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="ventas_mensuales_{año}_{mes:02d}.csv"'
        
        total_mes = ventas.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_mes = ventas.count()
        
        writer = csv.writer(response)
        writer.writerow(['Reporte de Ventas Mensuales'])
        writer.writerow(['Período', f'{año}-{mes:02d}'])
        writer.writerow(['Total Mensual', f"${float(total_mes):,.0f}"])
        writer.writerow(['Cantidad de Ventas', cantidad_mes])
        writer.writerow([])
        writer.writerow(['Código', 'Nombre Producto', 'Cantidad Vendida', 'Total Vendido'])
        
        for producto in top_productos:
            writer.writerow([
                producto.get('producto__codigo', ''),
                producto.get('producto__nombre', ''),
                producto.get('cantidad_vendida', 0),
                f"${float(producto.get('total_vendido', 0)):,.0f}"
            ])
        
        return response

    def _generar_csv_margen_productos(self, productos):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="margen_productos.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Código', 'Nombre', 'Costo', 'Precio Venta', 'Margen %', 'Ganancia Unitaria', 'Stock'])
        
        for p in productos:
            writer.writerow([
                p.codigo,
                p.nombre,
                p.costo,
                p.precio_venta,
                float(p.margen_calculado),
                float(p.ganancia_unitaria),
                p.stock_actual
            ])
        
        return response

    def _generar_csv_rotacion(self, productos):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="rotacion_inventario.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Código', 'Nombre', 'Stock Actual', 'Vendido 30 días', 'Rotación'])
        
        for p in productos:
            if p.stock_actual > 0:
                rotacion = (p.cantidad_vendida or 0) / p.stock_actual
                writer.writerow([
                    p.codigo,
                    p.nombre,
                    p.stock_actual,
                    p.cantidad_vendida or 0,
                    rotacion
                ])
        
        return response

    def _generar_csv_quiebres(self, quiebres, fecha_inicio, fecha_fin):
        """Método legacy - ahora usa Excel"""
        return self._generar_excel_quiebres(quiebres, fecha_inicio, fecha_fin)
    
    def _generar_excel_quiebres(self, quiebres, fecha_inicio, fecha_fin):
        """Generar Excel de quiebres de stock con diseño profesional"""
        if not OPENPYXL_AVAILABLE:
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="quiebres_semana_{fecha_inicio}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Reporte de Quiebres de Stock', f'{fecha_inicio} a {fecha_fin}'])
            writer.writerow([])
            writer.writerow(['Código', 'Nombre Producto', 'Cantidad de Quiebres', 'Fecha Primer Quiebre', 'Stock Mínimo Alcanzado'])
            for quiebre in quiebres:
                writer.writerow([
                    quiebre.get('producto__codigo', ''),
                    quiebre.get('producto__nombre', ''),
                    quiebre.get('cantidad_quiebres', 0),
                    quiebre.get('fecha_primer_quiebre', ''),
                    quiebre.get('stock_minimo_alcanzado', 0)
                ])
            return response
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Quiebres Stock"
        
        estilos = self._obtener_estilos_excel('1976D2')
        if not estilos:
            return self._generar_csv_quiebres(quiebres, fecha_inicio, fecha_fin)
        
        row_start = self._aplicar_titulo_excel(
            ws, 
            f'REPORTE DE QUIEBRES DE STOCK',
            f'Período: {fecha_inicio} a {fecha_fin} | Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de quiebres: {quiebres.count()}',
            5,
            estilos
        )
        
        headers = ['Código', 'Nombre Producto', 'Cantidad de Quiebres', 'Fecha Primer Quiebre', 'Stock Mínimo Alcanzado']
        self._aplicar_headers_excel(ws, headers, row_start, estilos)
        
        for idx, quiebre in enumerate(quiebres, row_start + 1):
            # Código
            cell = ws.cell(row=idx, column=1, value=quiebre.get('producto__codigo', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Nombre
            cell = ws.cell(row=idx, column=2, value=quiebre.get('producto__nombre', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Cantidad de Quiebres
            cantidad = quiebre.get('cantidad_quiebres', 0)
            cell = ws.cell(row=idx, column=3, value=cantidad)
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            # Fecha Primer Quiebre
            fecha_str = ''
            if quiebre.get('fecha_primer_quiebre'):
                try:
                    fecha_obj = quiebre.get('fecha_primer_quiebre')
                    if hasattr(fecha_obj, 'strftime'):
                        fecha_str = fecha_obj.strftime('%d/%m/%Y %H:%M:%S')
                    else:
                        fecha_str = str(fecha_obj)
                except:
                    fecha_str = str(quiebre.get('fecha_primer_quiebre', ''))
            cell = ws.cell(row=idx, column=4, value=fecha_str)
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['center_align']
            
            # Stock Mínimo Alcanzado
            stock_min = quiebre.get('stock_minimo_alcanzado', 0)
            cell = ws.cell(row=idx, column=5, value=stock_min)
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            if stock_min <= 0:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            
            ws.row_dimensions[idx].height = 18
        
        column_widths = [15, 35, 20, 20, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'quiebres_semana_{fecha_inicio.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def _generar_csv_compras_mensuales(self, compras, top_productos, compras_por_proveedor, año, mes):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="compras_mensuales_{año}_{mes:02d}.csv"'
        
        total_mes = compras.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_mes = compras.count()
        
        writer = csv.writer(response)
        writer.writerow(['Reporte de Compras Mensuales'])
        writer.writerow(['Período', f'{año}-{mes:02d}'])
        writer.writerow(['Total Mensual', f"${float(total_mes):,.0f}"])
        writer.writerow(['Cantidad de Compras', cantidad_mes])
        writer.writerow([])
        writer.writerow(['Top Productos Comprados'])
        writer.writerow(['Código', 'Nombre Producto', 'Cantidad Comprada', 'Total Comprado'])
        
        for producto in top_productos:
            writer.writerow([
                producto.get('producto__codigo', ''),
                producto.get('producto__nombre', ''),
                producto.get('cantidad_comprada', 0),
                f"${float(producto.get('total_comprado', 0)):,.0f}"
            ])
        
        writer.writerow([])
        writer.writerow(['Compras por Proveedor'])
        writer.writerow(['Proveedor', 'Total Comprado', 'Cantidad Compras'])
        
        for proveedor in compras_por_proveedor:
            writer.writerow([
                proveedor.get('proveedor__nombre', ''),
                f"${float(proveedor.get('total_comprado', 0)):,.0f}",
                proveedor.get('cantidad_compras', 0)
            ])
        
        return response

    def _generar_csv_proveedores(self, proveedores):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="reporte_proveedores.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Reporte de Proveedores'])
        writer.writerow(['Fecha', timezone.now().date().isoformat()])
        writer.writerow([])
        writer.writerow(['Nombre', 'RUT', 'Contacto', 'Teléfono', 'Email', 'Cantidad Productos', 'Total Compras 30 días', 'Cantidad Compras 30 días'])
        
        for p in proveedores:
            writer.writerow([
                p.nombre,
                p.rut or '',
                p.contacto or '',
                p.telefono or '',
                p.email or '',
                p.cantidad_productos or 0,
                f"${float(p.total_compras or 0):,.0f}",
                p.cantidad_compras or 0
            ])
        
        return response

    def _generar_csv_productos(self, productos, productos_bajo_stock, productos_mas_vendidos):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="reporte_productos.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Reporte de Productos'])
        writer.writerow(['Fecha', timezone.now().date().isoformat()])
        writer.writerow([])
        writer.writerow(['Productos con Stock Bajo'])
        writer.writerow(['Código', 'Nombre', 'Stock Actual', 'Stock Mínimo', 'Categoría'])
        
        for producto in productos_bajo_stock:
            writer.writerow([
                producto.get('codigo', ''),
                producto.get('nombre', ''),
                producto.get('stock_actual', 0),
                producto.get('stock_minimo', 0),
                producto.get('categoria__nombre', '')
            ])
        
        writer.writerow([])
        writer.writerow(['Productos Más Vendidos (30 días)'])
        writer.writerow(['Código', 'Nombre', 'Cantidad Vendida', 'Precio Venta'])
        
        # Convertir QuerySet a lista de diccionarios si es necesario
        if hasattr(productos_mas_vendidos, 'values'):
            productos_mas_vendidos_list = list(productos_mas_vendidos.values(
                'codigo', 'nombre', 'cantidad_vendida_30dias', 'precio_venta'
            ))
        else:
            productos_mas_vendidos_list = productos_mas_vendidos
        
        for producto in productos_mas_vendidos_list:
            writer.writerow([
                producto.get('codigo', ''),
                producto.get('nombre', ''),
                producto.get('cantidad_vendida_30dias', 0),
                f"${float(producto.get('precio_venta', 0)):,.0f}"
            ])
        
        return response

    # ========== MÉTODOS PARA GENERAR EXCEL CON DISEÑO PROFESIONAL ==========
    
    def _obtener_estilos_excel(self, color_tema='1976D2'):
        """Obtener estilos estándar para Excel (mismo diseño que Historial de Movimientos)"""
        if not OPENPYXL_AVAILABLE:
            return None
        
        # Colores del tema (azul por defecto, como historial de movimientos)
        title_fill = PatternFill(start_color=color_tema, end_color=color_tema, fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        # Header usa un azul más claro para contraste
        header_color = "2196F3" if color_tema == "1976D2" else "2e7d32"
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        data_font = Font(size=10)
        number_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        return {
            'title_fill': title_fill,
            'title_font': title_font,
            'header_fill': header_fill,
            'header_font': header_font,
            'border_style': border_style,
            'data_font': data_font,
            'number_font': number_font,
            'center_align': center_align,
            'left_align': left_align,
            'right_align': right_align,
        }
    
    def _aplicar_titulo_excel(self, ws, titulo, info_adicional, num_cols, estilos):
        """Aplicar título e información a una hoja Excel"""
        if not OPENPYXL_AVAILABLE or not estilos:
            return 2
        
        # Título del reporte
        ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        title_cell = ws['A1']
        title_cell.value = titulo
        title_cell.fill = estilos['title_fill']
        title_cell.font = estilos['title_font']
        title_cell.alignment = estilos['center_align']
        title_cell.border = estilos['border_style']
        ws.row_dimensions[1].height = 25
        
        # Información del reporte
        if info_adicional:
            ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
            info_cell = ws['A2']
            info_cell.value = info_adicional
            info_cell.font = Font(size=9, italic=True, color='666666')
            info_cell.alignment = estilos['center_align']
            ws.row_dimensions[2].height = 18
            return 3  # Retornar la fila donde empiezan los headers
        return 2
    
    def _aplicar_headers_excel(self, ws, headers, row_num, estilos):
        """Aplicar encabezados a una hoja Excel"""
        if not OPENPYXL_AVAILABLE or not estilos:
            return
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.fill = estilos['header_fill']
            cell.font = estilos['header_font']
            cell.alignment = estilos['center_align']
            cell.border = estilos['border_style']
        ws.row_dimensions[row_num].height = 20
    
    def _generar_excel_ventas_mensuales(self, ventas, top_productos, año, mes):
        """Generar Excel de ventas mensuales con diseño profesional"""
        if not OPENPYXL_AVAILABLE:
            return self._generar_csv_ventas_mensuales(ventas, top_productos, año, mes)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas Mensuales"
        
        total_mes = ventas.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_mes = ventas.count()
        meses_nombres = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        
        # Obtener estilos estándar
        estilos = self._obtener_estilos_excel('1976D2')
        if not estilos:
            return self._generar_csv_ventas_mensuales(ventas, top_productos, año, mes)
        
        # Título y resumen
        row_start = self._aplicar_titulo_excel(
            ws, 
            f'REPORTE DE VENTAS MENSUALES - {meses_nombres[mes]} {año}',
            f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Total: ${float(total_mes):,.0f} | Cantidad de ventas: {cantidad_mes}',
            5,
            estilos
        )
        
        # Encabezados
        headers = ['#', 'Código', 'Nombre Producto', 'Cantidad Vendida', 'Total Vendido']
        self._aplicar_headers_excel(ws, headers, row_start, estilos)
        
        # Datos
        for idx, producto in enumerate(top_productos, row_start + 1):
            # #
            cell = ws.cell(row=idx, column=1, value=idx - row_start)
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['center_align']
            
            # Código
            cell = ws.cell(row=idx, column=2, value=producto.get('producto__codigo', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Nombre
            cell = ws.cell(row=idx, column=3, value=producto.get('producto__nombre', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Cantidad
            cell = ws.cell(row=idx, column=4, value=producto.get('cantidad_vendida', 0))
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['right_align']
            
            # Total
            total = float(producto.get('total_vendido', 0))
            cell = ws.cell(row=idx, column=5, value=f"${total:,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            ws.row_dimensions[idx].height = 18
        
        # Ajustar columnas
        column_widths = [6, 15, 35, 18, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'ventas_mensuales_{año}_{mes:02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    def _generar_excel_compras_mensuales(self, compras, top_productos, compras_por_proveedor, año, mes):
        """Generar Excel de compras mensuales con diseño profesional"""
        if not OPENPYXL_AVAILABLE:
            return self._generar_csv_compras_mensuales(compras, top_productos, compras_por_proveedor, año, mes)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Compras Mensuales"
        
        total_mes = compras.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        cantidad_mes = compras.count()
        meses_nombres = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        
        # Obtener estilos estándar
        estilos = self._obtener_estilos_excel('1976D2')
        if not estilos:
            return self._generar_csv_compras_mensuales(compras, top_productos, compras_por_proveedor, año, mes)
        
        # Título y resumen
        row_start = self._aplicar_titulo_excel(
            ws, 
            f'REPORTE DE COMPRAS MENSUALES - {meses_nombres[mes]} {año}',
            f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Total: ${float(total_mes):,.0f} | Cantidad de compras: {cantidad_mes}',
            5,
            estilos
        )
        
        # Sección: Top Productos Comprados
        ws.merge_cells(f'A{row_start}:E{row_start}')
        section_cell = ws[f'A{row_start}']
        section_cell.value = 'TOP 10 PRODUCTOS COMPRADOS'
        section_cell.fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
        section_cell.font = Font(bold=True, color="FFFFFF", size=12)
        section_cell.alignment = estilos['center_align']
        section_cell.border = estilos['border_style']
        ws.row_dimensions[row_start].height = 22
        row_start += 1
        
        # Encabezados productos
        headers = ['#', 'Código', 'Nombre Producto', 'Cantidad Comprada', 'Total Comprado']
        self._aplicar_headers_excel(ws, headers, row_start, estilos)
        row_start += 1
        
        # Datos productos
        if top_productos:
            for idx, producto in enumerate(top_productos, row_start):
                # #
                cell = ws.cell(row=idx, column=1, value=idx - row_start + 1)
                cell.border = estilos['border_style']
                cell.font = estilos['number_font']
                cell.alignment = estilos['center_align']
                
                # Código
                cell = ws.cell(row=idx, column=2, value=producto.get('producto__codigo', ''))
                cell.border = estilos['border_style']
                cell.font = estilos['data_font']
                cell.alignment = estilos['left_align']
                
                # Nombre
                cell = ws.cell(row=idx, column=3, value=producto.get('producto__nombre', ''))
                cell.border = estilos['border_style']
                cell.font = estilos['data_font']
                cell.alignment = estilos['left_align']
                
                # Cantidad
                cell = ws.cell(row=idx, column=4, value=producto.get('cantidad_comprada', 0))
                cell.border = estilos['border_style']
                cell.font = estilos['number_font']
                cell.alignment = estilos['right_align']
                
                # Total
                total = float(producto.get('total_comprado', 0))
                cell = ws.cell(row=idx, column=5, value=f"${total:,.0f}")
                cell.border = estilos['border_style']
                cell.font = Font(size=10, bold=True)
                cell.alignment = estilos['right_align']
                
                ws.row_dimensions[idx].height = 18
            
            # Calcular la siguiente fila después de los productos
            row_start = row_start + len(top_productos) + 1
        else:
            # No hay productos, dejar un espacio después de los headers
            row_start = row_start + 1
        
        # Sección: Compras por Proveedor
        ws.merge_cells(f'A{row_start}:C{row_start}')
        section_cell2 = ws[f'A{row_start}']
        section_cell2.value = 'COMPRAS POR PROVEEDOR'
        section_cell2.fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
        section_cell2.font = Font(bold=True, color="FFFFFF", size=12)
        section_cell2.alignment = estilos['center_align']
        section_cell2.border = estilos['border_style']
        ws.row_dimensions[row_start].height = 22
        row_start += 1
        
        # Encabezados proveedores
        headers_prov = ['Proveedor', 'Total Comprado', 'Cantidad Compras']
        self._aplicar_headers_excel(ws, headers_prov, row_start, estilos)
        row_start += 1
        
        # Datos proveedores
        for idx, proveedor in enumerate(compras_por_proveedor, row_start):
            # Proveedor
            cell = ws.cell(row=idx, column=1, value=proveedor.get('proveedor__nombre', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Total
            total = float(proveedor.get('total_comprado', 0))
            cell = ws.cell(row=idx, column=2, value=f"${total:,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            # Cantidad
            cell = ws.cell(row=idx, column=3, value=proveedor.get('cantidad_compras', 0))
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['center_align']
            
            ws.row_dimensions[idx].height = 18
        
        # Ajustar columnas
        column_widths = [6, 15, 35, 18, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta - usar el mismo método que historial de pedidos
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'compras_mensuales_{año}_{mes:02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar directamente en la respuesta
        wb.save(response)
        return response
    
    def _generar_excel_proveedores(self, proveedores):
        """Generar Excel de proveedores con diseño profesional"""
        if not OPENPYXL_AVAILABLE:
            return self._generar_csv_proveedores(proveedores)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Proveedores"
        
        # Obtener estilos estándar
        estilos = self._obtener_estilos_excel('1976D2')
        if not estilos:
            return self._generar_csv_proveedores(proveedores)
        
        # Título y resumen
        row_start = self._aplicar_titulo_excel(
            ws, 
            'REPORTE DE PROVEEDORES',
            f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de proveedores activos: {proveedores.count()}',
            8,
            estilos
        )
        
        # Encabezados
        headers = ['Nombre', 'RUT', 'Contacto', 'Teléfono', 'Email', 'Cant. Productos', 'Total Compras 30 días', 'Cant. Compras 30 días']
        self._aplicar_headers_excel(ws, headers, row_start, estilos)
        
        # Datos
        for idx, p in enumerate(proveedores, row_start + 1):
            # Nombre
            cell = ws.cell(row=idx, column=1, value=p.nombre)
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # RUT
            cell = ws.cell(row=idx, column=2, value=p.rut or '')
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Contacto
            cell = ws.cell(row=idx, column=3, value=p.contacto or '')
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Teléfono
            cell = ws.cell(row=idx, column=4, value=p.telefono or '')
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Email
            cell = ws.cell(row=idx, column=5, value=p.email or '')
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Cant. Productos
            cell = ws.cell(row=idx, column=6, value=p.cantidad_productos or 0)
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['center_align']
            
            # Total Compras
            total = float(p.total_compras or 0)
            cell = ws.cell(row=idx, column=7, value=f"${total:,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            # Cant. Compras
            cell = ws.cell(row=idx, column=8, value=p.cantidad_compras or 0)
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['center_align']
            
            ws.row_dimensions[idx].height = 18
        
        # Ajustar columnas
        column_widths = [28, 15, 20, 15, 30, 15, 22, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'reporte_proveedores_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    def _generar_excel_productos(self, productos, productos_bajo_stock, productos_mas_vendidos):
        """Generar Excel de productos con diseño profesional"""
        if not OPENPYXL_AVAILABLE:
            return self._generar_csv_productos(productos, productos_bajo_stock, productos_mas_vendidos)
        
        wb = Workbook()
        
        # Obtener estilos estándar
        estilos = self._obtener_estilos_excel('1976D2')
        if not estilos:
            return self._generar_csv_productos(productos, productos_bajo_stock, productos_mas_vendidos)
        
        # Hoja 1: Productos con Stock Bajo
        ws1 = wb.active
        ws1.title = "Stock Bajo"
        
        # Convertir QuerySet a lista de diccionarios si es necesario
        if hasattr(productos_bajo_stock, 'values'):
            productos_bajo_stock_list = list(productos_bajo_stock.values(
                'codigo', 'nombre', 'stock_actual', 'stock_minimo', 'categoria__nombre'
            ))
        else:
            productos_bajo_stock_list = productos_bajo_stock
        
        row_start = self._aplicar_titulo_excel(
            ws1, 
            'PRODUCTOS CON STOCK BAJO',
            f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Total: {len(productos_bajo_stock_list)} productos',
            5,
            estilos
        )
        
        headers = ['Código', 'Nombre', 'Categoría', 'Stock Actual', 'Stock Mínimo']
        self._aplicar_headers_excel(ws1, headers, row_start, estilos)
        
        for idx, producto in enumerate(productos_bajo_stock_list, row_start + 1):
            # Código
            cell = ws1.cell(row=idx, column=1, value=producto.get('codigo', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Nombre
            cell = ws1.cell(row=idx, column=2, value=producto.get('nombre', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Categoría
            cell = ws1.cell(row=idx, column=3, value=producto.get('categoria__nombre', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Stock Actual con color
            stock_actual = producto.get('stock_actual', 0)
            cell = ws1.cell(row=idx, column=4, value=stock_actual)
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            if stock_actual <= 0:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            
            # Stock Mínimo
            cell = ws1.cell(row=idx, column=5, value=producto.get('stock_minimo', 0))
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['right_align']
            
            ws1.row_dimensions[idx].height = 18
        
        column_widths1 = [15, 35, 20, 15, 15]
        for col_num, width in enumerate(column_widths1, 1):
            ws1.column_dimensions[get_column_letter(col_num)].width = width
        
        # Hoja 2: Productos Más Vendidos
        ws2 = wb.create_sheet(title="Más Vendidos")
        
        # Convertir QuerySet a lista de diccionarios si es necesario
        if hasattr(productos_mas_vendidos, 'values'):
            productos_mas_vendidos_list = list(productos_mas_vendidos.values(
                'codigo', 'nombre', 'cantidad_vendida_30dias', 'precio_venta'
            ))
        else:
            productos_mas_vendidos_list = productos_mas_vendidos
        
        row_start2 = self._aplicar_titulo_excel(
            ws2, 
            'PRODUCTOS MÁS VENDIDOS (Últimos 30 días)',
            f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Total: {len(productos_mas_vendidos_list)} productos',
            5,
            estilos
        )
        
        headers2 = ['#', 'Código', 'Nombre', 'Cantidad Vendida', 'Precio Venta']
        self._aplicar_headers_excel(ws2, headers2, row_start2, estilos)
        
        for idx, producto in enumerate(productos_mas_vendidos_list, row_start2 + 1):
            # #
            cell = ws2.cell(row=idx, column=1, value=idx - row_start2)
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['center_align']
            
            # Código
            cell = ws2.cell(row=idx, column=2, value=producto.get('codigo', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Nombre
            cell = ws2.cell(row=idx, column=3, value=producto.get('nombre', ''))
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Cantidad
            cell = ws2.cell(row=idx, column=4, value=producto.get('cantidad_vendida_30dias', 0))
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            # Precio
            precio = float(producto.get('precio_venta', 0))
            cell = ws2.cell(row=idx, column=5, value=f"${precio:,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            ws2.row_dimensions[idx].height = 18
        
        column_widths2 = [6, 15, 35, 18, 18]
        for col_num, width in enumerate(column_widths2, 1):
            ws2.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'reporte_productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    def _generar_excel_margen_productos(self, productos):
        """Generar Excel de margen de productos con diseño profesional"""
        if not OPENPYXL_AVAILABLE:
            return self._generar_csv_margen_productos(productos)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Margen Productos"
        
        # Obtener estilos estándar
        estilos = self._obtener_estilos_excel('1976D2')
        if not estilos:
            return self._generar_csv_margen_productos(productos)
        
        row_start = self._aplicar_titulo_excel(
            ws, 
            'REPORTE DE MARGEN DE GANANCIA POR PRODUCTO',
            f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de productos: {productos.count()}',
            7,
            estilos
        )
        
        headers = ['Código', 'Nombre', 'Costo', 'Precio Venta', 'Margen %', 'Ganancia Unitaria', 'Stock']
        self._aplicar_headers_excel(ws, headers, row_start, estilos)
        
        for idx, p in enumerate(productos, row_start + 1):
            # Código
            cell = ws.cell(row=idx, column=1, value=p.codigo)
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Nombre
            cell = ws.cell(row=idx, column=2, value=p.nombre)
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Costo
            cell = ws.cell(row=idx, column=3, value=f"${float(p.costo):,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            # Precio Venta
            cell = ws.cell(row=idx, column=4, value=f"${float(p.precio_venta):,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            # Margen % con color
            margen = float(p.margen_calculado)
            cell = ws.cell(row=idx, column=5, value=f"{margen:.2f}%")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            # Color según margen
            if margen > 50:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif margen > 30:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif margen > 0:
                cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            # Ganancia Unitaria
            cell = ws.cell(row=idx, column=6, value=f"${float(p.ganancia_unitaria):,.0f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            
            # Stock
            cell = ws.cell(row=idx, column=7, value=p.stock_actual)
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['right_align']
            
            ws.row_dimensions[idx].height = 18
        
        column_widths = [15, 35, 15, 15, 12, 18, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'margen_productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    def _generar_excel_rotacion(self, productos):
        """Generar Excel de rotación de inventario con diseño profesional"""
        if not OPENPYXL_AVAILABLE:
            return self._generar_csv_rotacion(productos)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Rotación Inventario"
        
        # Obtener estilos estándar
        estilos = self._obtener_estilos_excel('1976D2')
        if not estilos:
            return self._generar_csv_rotacion(productos)
        
        productos_con_rotacion = [p for p in productos if p.stock_actual > 0]
        
        row_start = self._aplicar_titulo_excel(
            ws, 
            'REPORTE DE ROTACIÓN DE INVENTARIO',
            f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")} | Productos analizados: {len(productos_con_rotacion)}',
            5,
            estilos
        )
        
        headers = ['Código', 'Nombre', 'Stock Actual', 'Vendido 30 días', 'Rotación']
        self._aplicar_headers_excel(ws, headers, row_start, estilos)
        
        for idx, p in enumerate(productos_con_rotacion, row_start + 1):
            # Código
            cell = ws.cell(row=idx, column=1, value=p.codigo)
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Nombre
            cell = ws.cell(row=idx, column=2, value=p.nombre)
            cell.border = estilos['border_style']
            cell.font = estilos['data_font']
            cell.alignment = estilos['left_align']
            
            # Stock Actual
            cell = ws.cell(row=idx, column=3, value=p.stock_actual)
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['right_align']
            
            # Vendido 30 días
            vendido = p.cantidad_vendida or 0
            cell = ws.cell(row=idx, column=4, value=vendido)
            cell.border = estilos['border_style']
            cell.font = estilos['number_font']
            cell.alignment = estilos['right_align']
            
            # Rotación con color
            rotacion = vendido / p.stock_actual if p.stock_actual > 0 else 0
            cell = ws.cell(row=idx, column=5, value=f"{rotacion:.2f}")
            cell.border = estilos['border_style']
            cell.font = Font(size=10, bold=True)
            cell.alignment = estilos['right_align']
            # Color según rotación
            if rotacion > 1:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif rotacion > 0.5:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif rotacion > 0:
                cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            ws.row_dimensions[idx].height = 18
        
        column_widths = [15, 35, 15, 18, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'rotacion_inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

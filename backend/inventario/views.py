from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.db.models import F
from django.utils import timezone
from django.core.cache import cache
from django.views.decorators.cache import cache_page
from django.utils.decorators import method_decorator
from usuarios.permissions import PuedeProductos, EsAdministradorOReadOnly
from .models import Proveedor, Categoria, Producto, MovimientoStock, PedidoProveedor
from .serializers import (
    ProveedorSerializer,
    CategoriaSerializer,
    ProductoSerializer,
    MovimientoStockSerializer,
    PedidoProveedorSerializer
)


class ProveedorViewSet(viewsets.ModelViewSet):
    queryset = Proveedor.objects.filter(activo=True)
    serializer_class = ProveedorSerializer
    search_fields = ['nombre', 'rut', 'contacto']
    permission_classes = [IsAuthenticated, EsAdministradorOReadOnly]  # Lectura para todos, escritura solo admin

    @action(detail=True, methods=['post'])
    def enviar_pedido(self, request, pk=None):
        """Enviar correo de pedido a un proveedor"""
        from django.core.mail import send_mail
        from django.template.loader import render_to_string
        from django.conf import settings
        from datetime import datetime
        from usuarios.models import Configuracion
        
        proveedor = self.get_object()
        
        # Obtener dirección del minimarket
        direccion_minimarket = Configuracion.obtener_valor('direccion_minimarket', '')
        
        if not proveedor.email:
            return Response(
                {'error': 'El proveedor no tiene un correo electrónico registrado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        items = request.data.get('items', [])
        notas = request.data.get('notas', '')
        fecha_estimada = request.data.get('fecha_estimada', '')
        
        if not items or len(items) == 0:
            return Response(
                {'error': 'Debe incluir al menos un producto en el pedido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obtener información de los productos
        productos_info = []
        for item in items:
            try:
                producto = Producto.objects.get(id=item.get('producto'))
                cantidad = item.get('cantidad', 1)
                productos_info.append({
                    'codigo': producto.codigo,
                    'nombre': producto.nombre,
                    'cantidad': cantidad,
                    'unidad_medida': producto.unidad_medida,
                })
            except Producto.DoesNotExist:
                continue
        
        if not productos_info:
            return Response(
                {'error': 'No se encontraron productos válidos'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Formatear fecha estimada
        fecha_estimada_str = ''
        if fecha_estimada:
            try:
                fecha_obj = datetime.strptime(fecha_estimada, '%Y-%m-%d')
                fecha_estimada_str = fecha_obj.strftime('%d/%m/%Y')
            except:
                fecha_estimada_str = fecha_estimada
        
        # Calcular total de items
        total_items = sum(item['cantidad'] for item in productos_info)
        
        # Crear el contenido del correo
        asunto = f'Solicitud de Pedido - Minimarket La Esquina'
        
        mensaje_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; line-height: 1.6; color: #333333; margin: 0; padding: 0; background-color: #f5f5f5;">
            <div style="max-width: 650px; margin: 20px auto; background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <!-- Header -->
                <div style="background: linear-gradient(135deg, #2e7d32 0%, #1b5e20 100%); color: white; padding: 30px 40px; text-align: center;">
                    <h1 style="margin: 0; font-size: 24px; font-weight: 600;">SOLICITUD DE PEDIDO</h1>
                    <p style="margin: 10px 0 0 0; font-size: 14px; opacity: 0.9;">Minimarket La Esquina</p>
                </div>
                
                <!-- Contenido Principal -->
                <div style="padding: 40px;">
                    <p style="margin: 0 0 20px 0; font-size: 16px;">
                        Estimado/a <strong>{proveedor.contacto or proveedor.nombre}</strong>,
                    </p>
                    
                    <p style="margin: 0 0 30px 0; font-size: 15px; color: #555555;">
                        Por medio de la presente, nos dirigimos a usted para solicitar formalmente el siguiente pedido de productos. 
                        Agradecemos su atención y esperamos su confirmación de disponibilidad y condiciones de entrega.
                    </p>
                    
                    <!-- Resumen del Pedido -->
                    <div style="background-color: #f8f9fa; border-left: 4px solid #2e7d32; padding: 15px 20px; margin-bottom: 30px; border-radius: 4px;">
                        <p style="margin: 0; font-size: 14px; color: #666666;">
                            <strong>Resumen del Pedido:</strong><br>
                            • Total de productos: <strong>{len(productos_info)}</strong><br>
                            • Total de unidades: <strong>{total_items}</strong>
                        </p>
                    </div>
                    
                    <!-- Tabla de Productos -->
                    <div style="margin-bottom: 30px;">
                        <h3 style="margin: 0 0 15px 0; font-size: 18px; color: #2e7d32; font-weight: 600;">Detalle de Productos Solicitados</h3>
                        <table style="width: 100%; border-collapse: collapse; background-color: #ffffff; border-radius: 6px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
                            <thead>
                                <tr style="background-color: #2e7d32; color: white;">
                                    <th style="padding: 12px 15px; text-align: left; font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">#</th>
                                    <th style="padding: 12px 15px; text-align: left; font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Descripción del Producto</th>
                                    <th style="padding: 12px 15px; text-align: right; font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Cantidad</th>
                                </tr>
                            </thead>
                            <tbody>
        """
        
        for idx, producto in enumerate(productos_info, 1):
            mensaje_html += f"""
                                <tr style="border-bottom: 1px solid #e9ecef;">
                                    <td style="padding: 12px 15px; font-size: 14px; color: #666666; font-weight: 500;">{idx}</td>
                                    <td style="padding: 12px 15px; font-size: 14px;">
                                        <strong style="color: #333333;">{producto['nombre']}</strong><br>
                                        <span style="color: #999999; font-size: 12px;">Ref: {producto['codigo']}</span>
                                    </td>
                                    <td style="padding: 12px 15px; text-align: right; font-size: 14px; font-weight: 600; color: #2e7d32;">{producto['cantidad']} {producto['unidad_medida']}</td>
                                </tr>
            """
        
        mensaje_html += """
                            </tbody>
                        </table>
                    </div>
                    
                    <!-- Información Adicional -->
        """
        
        if fecha_estimada_str:
            mensaje_html += f"""
                    <div style="background-color: #fff3cd; border-left: 4px solid #ffc107; padding: 15px 20px; margin-bottom: 20px; border-radius: 4px;">
                        <p style="margin: 0; font-size: 14px; color: #856404;">
                            <strong>📅 Fecha Estimada de Entrega:</strong> {fecha_estimada_str}
                        </p>
                    </div>
            """
        
        if notas:
            mensaje_html += f"""
                    <div style="background-color: #e7f3ff; border-left: 4px solid #2196f3; padding: 15px 20px; margin-bottom: 20px; border-radius: 4px;">
                        <p style="margin: 0 0 8px 0; font-size: 14px; font-weight: 600; color: #0d47a1;">Notas Adicionales:</p>
                        <p style="margin: 0; font-size: 14px; color: #1565c0; white-space: pre-line;">{notas}</p>
                    </div>
            """
        
        if direccion_minimarket:
            mensaje_html += f"""
                    <div style="background-color: #f0f4f8; border-left: 4px solid #1976d2; padding: 15px 20px; margin-bottom: 20px; border-radius: 4px;">
                        <p style="margin: 0 0 8px 0; font-size: 14px; font-weight: 600; color: #0d47a1;">📍 Dirección de Entrega:</p>
                        <p style="margin: 0; font-size: 14px; color: #1565c0; line-height: 1.8;">{direccion_minimarket.replace(chr(10), '<br>')}</p>
                    </div>
            """
        
        mensaje_html += """
                    <!-- Cierre -->
                    <div style="margin-top: 30px; padding-top: 25px; border-top: 2px solid #e9ecef;">
                        <p style="margin: 0 0 15px 0; font-size: 15px; color: #555555;">
                            Solicitamos de manera especial que nos confirme la disponibilidad de los productos solicitados, 
                            así como las condiciones de entrega, tiempos estimados y cualquier información relevante para 
                            coordinar la recepción del pedido.
                        </p>
                        <p style="margin: 0; font-size: 15px; color: #555555;">
                            Quedamos atentos a su respuesta y agradecemos de antemano su atención y colaboración.
                        </p>
                    </div>
                </div>
                
                <!-- Footer -->
                <div style="background-color: #f8f9fa; padding: 25px 40px; text-align: center; border-top: 1px solid #e9ecef;">
                    <p style="margin: 0 0 8px 0; font-size: 14px; font-weight: 600; color: #333333;">Atentamente,</p>
                    <p style="margin: 0; font-size: 14px; color: #2e7d32; font-weight: 600;">Equipo de Minimarket La Esquina</p>
                    <p style="margin: 10px 0 0 0; font-size: 12px; color: #999999;">
                        Este es un correo automático generado por nuestro sistema de gestión.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        mensaje_texto = f"""
═══════════════════════════════════════════════════════════════
           SOLICITUD DE PEDIDO - MINIMARKET LA ESQUINA
═══════════════════════════════════════════════════════════════

Estimado/a {proveedor.contacto or proveedor.nombre},

Por medio de la presente, nos dirigimos a usted para solicitar 
formalmente el siguiente pedido de productos. Agradecemos su 
atención y esperamos su confirmación de disponibilidad y 
condiciones de entrega.

───────────────────────────────────────────────────────────────
RESUMEN DEL PEDIDO
───────────────────────────────────────────────────────────────
Total de productos: {len(productos_info)}
Total de unidades: {total_items}

───────────────────────────────────────────────────────────────
DETALLE DE PRODUCTOS SOLICITADOS
───────────────────────────────────────────────────────────────
"""
        for idx, producto in enumerate(productos_info, 1):
            mensaje_texto += f"""
{idx}. {producto['nombre']}
   Referencia: {producto['codigo']}
   Cantidad: {producto['cantidad']} {producto['unidad_medida']}
"""
        
        mensaje_texto += "\n───────────────────────────────────────────────────────────────\n"
        
        if fecha_estimada_str:
            mensaje_texto += f"\n📅 FECHA ESTIMADA DE ENTREGA: {fecha_estimada_str}\n"
        
        if notas:
            mensaje_texto += f"\n📝 NOTAS ADICIONALES:\n{notas}\n"
        
        if direccion_minimarket:
            mensaje_texto += f"\n📍 DIRECCIÓN DE ENTREGA:\n{direccion_minimarket}\n"
        
        mensaje_texto += """
───────────────────────────────────────────────────────────────

Solicitamos de manera especial que nos confirme la disponibilidad 
de los productos solicitados, así como las condiciones de entrega, 
tiempos estimados y cualquier información relevante para coordinar 
la recepción del pedido.

Quedamos atentos a su respuesta y agradecemos de antemano su 
atención y colaboración.

───────────────────────────────────────────────────────────────

Atentamente,
Equipo de Minimarket La Esquina

───────────────────────────────────────────────────────────────
Este es un correo automático generado por nuestro sistema de gestión.
═══════════════════════════════════════════════════════════════
"""
        
        # Validar configuración de correo
        from_email = settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER
        if not from_email:
            return Response(
                {
                    'error': (
                        'Error de configuración de correo. Por favor, configura las siguientes variables en tu archivo .env:\n'
                        '1. EMAIL_HOST_USER=tu-email@gmail.com\n'
                        '2. EMAIL_HOST_PASSWORD=tu-contraseña-de-aplicación\n'
                        '3. DEFAULT_FROM_EMAIL=tu-email@gmail.com\n\n'
                        'Consulta el archivo CONFIGURACION_EMAIL.md para más detalles.'
                    )
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        try:
            send_mail(
                subject=asunto,
                message=mensaje_texto,
                from_email=from_email,
                recipient_list=[proveedor.email],
                html_message=mensaje_html,
                fail_silently=False,
            )
            
            # Guardar el pedido en el historial
            from .models import PedidoProveedor
            from datetime import datetime as dt
            
            fecha_estimada_obj = None
            if fecha_estimada:
                try:
                    fecha_estimada_obj = dt.strptime(fecha_estimada, '%Y-%m-%d').date()
                except:
                    pass
            
            total_items = sum(item['cantidad'] for item in productos_info)
            
            pedido = PedidoProveedor.objects.create(
                proveedor=proveedor,
                fecha_estimada_entrega=fecha_estimada_obj,
                notas=notas,
                email_enviado=proveedor.email,
                usuario=request.user.username if request.user.is_authenticated else 'Sistema',
                items=productos_info,
                total_items=total_items,
                estado='ENVIADO'
            )
            
            return Response({
                'mensaje': f'Correo enviado exitosamente a {proveedor.email}',
                'email_enviado': True,
                'destinatario': proveedor.email,
                'pedido_id': pedido.id
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            error_msg = str(e)
            # Mensajes más claros para errores comunes de Gmail
            if '535' in error_msg or 'BadCredentials' in error_msg or 'Username and Password not accepted' in error_msg:
                mensaje_error = (
                    'Error de autenticación con Gmail. Por favor, verifica:\n'
                    '1. Que estés usando una "Contraseña de aplicación" (App Password) y no tu contraseña normal\n'
                    '2. Que la contraseña de aplicación sea correcta\n'
                    '3. Que la verificación en dos pasos esté habilitada en tu cuenta de Gmail\n'
                    'Para generar una contraseña de aplicación: https://myaccount.google.com/apppasswords'
                )
            elif 'Connection refused' in error_msg or 'Network' in error_msg:
                mensaje_error = 'Error de conexión con el servidor de correo. Verifica tu conexión a internet.'
            elif 'timeout' in error_msg.lower():
                mensaje_error = 'Tiempo de espera agotado al conectar con el servidor de correo.'
            elif 'Invalid address' in error_msg or 'from_email' in error_msg.lower():
                mensaje_error = (
                    'Error de configuración de correo. El campo "from_email" está vacío.\n'
                    'Por favor, configura las siguientes variables en tu archivo .env:\n'
                    '1. EMAIL_HOST_USER=tu-email@gmail.com\n'
                    '2. EMAIL_HOST_PASSWORD=tu-contraseña-de-aplicación\n'
                    '3. DEFAULT_FROM_EMAIL=tu-email@gmail.com\n\n'
                    'Consulta el archivo CONFIGURACION_EMAIL.md para más detalles.'
                )
            else:
                mensaje_error = f'Error al enviar el correo: {error_msg}'
            
            return Response(
                {'error': mensaje_error},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def exportar_csv(self, request):
        """Exportar proveedores a Excel con diseño mejorado"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from datetime import datetime
        
        queryset = self.get_queryset()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Proveedores"
        
        # Estilos mejorados (mismo estilo que movimientos)
        title_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        data_font = Font(size=10)
        number_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Título del reporte
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = 'REPORTE DE PROVEEDORES'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.border = border_style
        ws.row_dimensions[1].height = 25
        
        # Información del reporte
        ws.merge_cells('A2:G2')
        info_cell = ws['A2']
        info_cell.value = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de proveedores: {queryset.count()}'
        info_cell.font = Font(size=9, italic=True, color='666666')
        info_cell.alignment = center_align
        ws.row_dimensions[2].height = 18
        
        # Encabezados
        headers = ['Nombre', 'RUT', 'Contacto', 'Teléfono', 'Email', 'Dirección', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        ws.row_dimensions[3].height = 20
        
        # Datos
        for row_num, proveedor in enumerate(queryset, 4):
            # Nombre
            cell = ws.cell(row=row_num, column=1, value=proveedor.nombre or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # RUT
            cell = ws.cell(row=row_num, column=2, value=proveedor.rut or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Contacto
            cell = ws.cell(row=row_num, column=3, value=proveedor.contacto or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Teléfono
            cell = ws.cell(row=row_num, column=4, value=proveedor.telefono or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Email
            cell = ws.cell(row=row_num, column=5, value=proveedor.email or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Dirección
            cell = ws.cell(row=row_num, column=6, value=proveedor.direccion or '')
            cell.border = border_style
            cell.font = Font(size=9)
            cell.alignment = left_align
            
            # Estado
            estado = 'Activo' if proveedor.activo else 'Inactivo'
            cell = ws.cell(row=row_num, column=7, value=estado)
            cell.border = border_style
            cell.font = data_font
            cell.alignment = center_align
            if proveedor.activo:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            ws.row_dimensions[row_num].height = 18
        
        # Ajustar ancho de columnas
        column_widths = [30, 15, 20, 15, 30, 40, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'proveedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.filter(activa=True)
    serializer_class = CategoriaSerializer
    search_fields = ['nombre']
    permission_classes = [IsAuthenticated, EsAdministradorOReadOnly]  # Lectura para todos, escritura solo admin


class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.select_related('categoria', 'proveedor').all()
    serializer_class = ProductoSerializer
    search_fields = ['codigo', 'nombre', 'codigo_barras']
    permission_classes = [IsAuthenticated, PuedeProductos]  # Solo administradores pueden gestionar productos

    def get_queryset(self):
        # Optimización: query optimizada con select_related y prefetch_related
        queryset = Producto.objects.select_related('categoria', 'proveedor').all()
        activo = self.request.query_params.get('activo', None)
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        return queryset
    
    def perform_create(self, serializer):
        """Invalidar caché al crear producto"""
        super().perform_create(serializer)
        cache.delete('productos_activos')
    
    def perform_update(self, serializer):
        """Invalidar caché al actualizar producto"""
        super().perform_update(serializer)
        cache.delete('productos_activos')
    
    def perform_destroy(self, instance):
        """Invalidar caché al eliminar producto"""
        super().perform_destroy(instance)
        cache.delete('productos_activos')

    def get_serializer_context(self):
        """Agregar request al contexto del serializer para generar URLs absolutas"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def destroy(self, request, *args, **kwargs):
        """Eliminar un producto con manejo de errores (soft delete)"""
        try:
            instance = self.get_object()
            
            # Verificar si el producto tiene ventas o compras asociadas
            from ventas.models import DetalleVenta
            from compras.models import DetalleCompra
            
            tiene_ventas = DetalleVenta.objects.filter(producto=instance).exists()
            tiene_compras = DetalleCompra.objects.filter(producto=instance).exists()
            
            if tiene_ventas or tiene_compras:
                # Si tiene ventas o compras, desactivar el producto en lugar de eliminarlo
                # Esto mantiene la integridad de los datos históricos
                cantidad_ventas = DetalleVenta.objects.filter(producto=instance).count()
                cantidad_compras = DetalleCompra.objects.filter(producto=instance).count()
                
                instance.activo = False
                instance.save()
                
                mensaje = f'El producto "{instance.nombre}" ha sido desactivado porque tiene '
                detalles = []
                if cantidad_ventas > 0:
                    detalles.append(f'{cantidad_ventas} venta(s)')
                if cantidad_compras > 0:
                    detalles.append(f'{cantidad_compras} compra(s)')
                
                mensaje += f'{", ".join(detalles)} asociada(s). '
                mensaje += 'Los datos históricos se mantienen intactos. Puedes reactivarlo editando el producto.'
                
                return Response(
                    {
                        'mensaje': mensaje,
                        'producto_desactivado': True,
                        'cantidad_ventas': cantidad_ventas,
                        'cantidad_compras': cantidad_compras
                    },
                    status=status.HTTP_200_OK
                )
            else:
                # Si no tiene ventas ni compras, eliminar físicamente
                # Eliminar la imagen si existe
                if instance.imagen:
                    try:
                        instance.imagen.delete(save=False)
                    except Exception:
                        pass  # Si falla la eliminación de la imagen, continuar
                
                self.perform_destroy(instance)
                return Response(
                    {
                        'mensaje': 'Producto eliminado exitosamente',
                        'producto_eliminado': True
                    },
                    status=status.HTTP_200_OK
                )
        except Exception as e:
            return Response(
                {'error': f'Error al eliminar el producto: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def ajustar_stock(self, request, pk=None):
        """Ajustar stock de un producto"""
        producto = self.get_object()
        cantidad = request.data.get('cantidad')
        motivo = request.data.get('motivo', 'Ajuste manual')
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'

        if cantidad is None:
            return Response(
                {'error': 'La cantidad es requerida'},
                status=status.HTTP_400_BAD_REQUEST
            )

        cantidad = int(cantidad)
        stock_anterior = producto.stock_actual
        stock_nuevo = stock_anterior + cantidad

        # Validar stock negativo
        if stock_nuevo < 0:
            return Response(
                {'error': 'No se permite stock negativo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            producto.stock_actual = stock_nuevo
            producto.save()

            MovimientoStock.objects.create(
                producto=producto,
                tipo='AJUSTE',
                cantidad=cantidad,
                stock_anterior=stock_anterior,
                stock_nuevo=stock_nuevo,
                motivo=motivo,
                usuario=usuario
            )

            # Verificar alertas de stock
            self._verificar_alerta_stock(producto, stock_anterior, stock_nuevo)
            
            # Invalidar caché de productos activos
            cache.delete('productos_activos')

        return Response({
            'mensaje': 'Stock ajustado correctamente',
            'stock_anterior': stock_anterior,
            'stock_nuevo': stock_nuevo
        })

    def _verificar_alerta_stock(self, producto, stock_anterior, stock_nuevo):
        """Verifica si debe crear o eliminar alertas de stock bajo"""
        try:
            from usuarios.models import AlertaStock
            
            # Si el stock anterior estaba por encima del mínimo y ahora está por debajo
            if stock_anterior > producto.stock_minimo and stock_nuevo <= producto.stock_minimo:
                # Crear nueva alerta solo si no hay una alerta sin leer
                if not AlertaStock.objects.filter(producto=producto, leida=False).exists():
                    AlertaStock.objects.create(producto=producto)
            # Si el stock estaba por debajo y ahora está por encima, marcar alertas como leídas
            elif stock_anterior <= producto.stock_minimo and stock_nuevo > producto.stock_minimo:
                AlertaStock.objects.filter(producto=producto, leida=False).update(
                    leida=True,
                    fecha_lectura=timezone.now()
                )
        except ImportError:
            pass

    @action(detail=False, methods=['get'])
    def exportar_csv(self, request):
        """Exportar productos a Excel con diseño mejorado"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from datetime import datetime
        
        # Obtener queryset con filtros aplicados
        queryset = self.get_queryset()
        
        # Aplicar filtros adicionales si existen
        categoria_id = request.query_params.get('categoria', None)
        proveedor_id = request.query_params.get('proveedor', None)
        
        if categoria_id:
            queryset = queryset.filter(categoria_id=categoria_id)
        if proveedor_id:
            queryset = queryset.filter(proveedor_id=proveedor_id)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Estilos mejorados (mismo estilo que movimientos)
        title_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        data_font = Font(size=10)
        number_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Título del reporte
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = 'REPORTE DE PRODUCTOS'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.border = border_style
        ws.row_dimensions[1].height = 25
        
        # Información del reporte
        ws.merge_cells('A2:J2')
        info_cell = ws['A2']
        info_cell.value = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de productos: {queryset.count()}'
        info_cell.font = Font(size=9, italic=True, color='666666')
        info_cell.alignment = center_align
        ws.row_dimensions[2].height = 18
        
        # Encabezados
        headers = ['Código', 'Nombre', 'Categoría', 'Proveedor', 'Stock Actual', 'Stock Mínimo', 'Precio Venta', 'Costo', 'Margen %', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        ws.row_dimensions[3].height = 20
        
        # Datos
        for row_num, producto in enumerate(queryset, 4):
            # Código
            cell = ws.cell(row=row_num, column=1, value=producto.codigo or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Nombre
            cell = ws.cell(row=row_num, column=2, value=producto.nombre or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Categoría
            cell = ws.cell(row=row_num, column=3, value=producto.categoria.nombre if producto.categoria else '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Proveedor
            cell = ws.cell(row=row_num, column=4, value=producto.proveedor.nombre if producto.proveedor else '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Stock Actual
            stock_actual = producto.stock_actual or 0
            cell = ws.cell(row=row_num, column=5, value=stock_actual)
            cell.border = border_style
            cell.font = Font(size=10, bold=True)
            cell.alignment = right_align
            # Color según stock
            if stock_actual <= 0:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            elif stock_actual <= producto.stock_minimo:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            
            # Stock Mínimo
            cell = ws.cell(row=row_num, column=6, value=producto.stock_minimo or 0)
            cell.border = border_style
            cell.font = number_font
            cell.alignment = right_align
            
            # Precio Venta
            cell = ws.cell(row=row_num, column=7, value=float(producto.precio_venta or 0))
            cell.border = border_style
            cell.font = Font(size=10, bold=True)
            cell.alignment = right_align
            cell.number_format = '#,##0'
            
            # Costo
            cell = ws.cell(row=row_num, column=8, value=float(producto.costo or 0))
            cell.border = border_style
            cell.font = number_font
            cell.alignment = right_align
            cell.number_format = '#,##0'
            
            # Margen %
            if producto.costo and producto.costo > 0:
                margen = ((producto.precio_venta or 0) - (producto.costo or 0)) / producto.costo * 100
            else:
                margen = 0
            cell = ws.cell(row=row_num, column=9, value=f"{margen:.2f}%")
            cell.border = border_style
            cell.font = Font(size=10, bold=True)
            cell.alignment = right_align
            # Color según margen
            if margen > 50:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif margen > 30:
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif margen <= 0:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            # Estado
            estado = 'Activo' if producto.activo else 'Inactivo'
            cell = ws.cell(row=row_num, column=10, value=estado)
            cell.border = border_style
            cell.font = data_font
            cell.alignment = center_align
            if producto.activo:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            ws.row_dimensions[row_num].height = 18
        
        # Ajustar ancho de columnas
        column_widths = [15, 30, 20, 25, 12, 12, 15, 15, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def bajo_stock(self, request):
        """Productos con stock bajo el mínimo - Accesible para todos los autenticados"""
        productos = Producto.objects.filter(
            activo=True,
            stock_actual__lte=F('stock_minimo')
        )
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)


class MovimientoStockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = MovimientoStock.objects.select_related('producto').all()
    serializer_class = MovimientoStockSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = MovimientoStock.objects.select_related('producto').all()
        
        # Filtro por producto
        producto_id = self.request.query_params.get('producto', None)
        if producto_id:
            queryset = queryset.filter(producto_id=producto_id)
        
        # Filtro por tipo
        tipo = self.request.query_params.get('tipo', None)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        
        # Filtro por usuario
        usuario = self.request.query_params.get('usuario', None)
        if usuario:
            queryset = queryset.filter(usuario__icontains=usuario)
        
        # Filtro por fecha desde
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        if fecha_desde:
            queryset = queryset.filter(fecha__gte=fecha_desde)
        
        # Filtro por fecha hasta
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)
        if fecha_hasta:
            queryset = queryset.filter(fecha__lte=fecha_hasta)
        
        return queryset.order_by('-fecha')
    
    @action(detail=False, methods=['get'])
    def exportar_csv(self, request):
        """Exportar movimientos a Excel con diseño mejorado"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from datetime import datetime
        
        queryset = self.get_queryset()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos de Stock"
        
        # Estilos mejorados
        title_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        data_font = Font(size=10)
        number_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Título del reporte
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = 'MOVIMIENTOS DE STOCK'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.border = border_style
        ws.row_dimensions[1].height = 25
        
        # Información del reporte
        ws.merge_cells('A2:I2')
        info_cell = ws['A2']
        info_cell.value = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de registros: {queryset.count()}'
        info_cell.font = Font(size=9, italic=True, color='666666')
        info_cell.alignment = center_align
        ws.row_dimensions[2].height = 18
        
        # Encabezados
        headers = ['Fecha', 'Producto', 'Código', 'Tipo', 'Cantidad', 'Stock Anterior', 'Stock Nuevo', 'Motivo', 'Usuario']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        ws.row_dimensions[3].height = 20
        
        # Datos
        for row_num, movimiento in enumerate(queryset, 4):
            # Fecha
            cell = ws.cell(row=row_num, column=1, value=movimiento.fecha.strftime('%d/%m/%Y %H:%M:%S'))
            cell.border = border_style
            cell.font = data_font
            cell.alignment = center_align
            
            # Producto
            cell = ws.cell(row=row_num, column=2, value=movimiento.producto.nombre if movimiento.producto else '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Código
            cell = ws.cell(row=row_num, column=3, value=movimiento.producto.codigo if movimiento.producto else '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Tipo con color
            tipo = movimiento.get_tipo_display()
            cell = ws.cell(row=row_num, column=4, value=tipo)
            cell.border = border_style
            cell.font = Font(size=10, bold=True)
            cell.alignment = center_align
            # Color según tipo
            if movimiento.tipo == 'ENTRADA':
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif movimiento.tipo == 'SALIDA':
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            elif movimiento.tipo == 'AJUSTE':
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            
            # Cantidad
            cell = ws.cell(row=row_num, column=5, value=movimiento.cantidad)
            cell.border = border_style
            cell.font = number_font
            cell.alignment = right_align
            
            # Stock Anterior
            cell = ws.cell(row=row_num, column=6, value=movimiento.stock_anterior)
            cell.border = border_style
            cell.font = number_font
            cell.alignment = right_align
            
            # Stock Nuevo
            cell = ws.cell(row=row_num, column=7, value=movimiento.stock_nuevo)
            cell.border = border_style
            cell.font = Font(size=10, bold=True)
            cell.alignment = right_align
            # Color según si aumentó o disminuyó
            if movimiento.stock_nuevo > movimiento.stock_anterior:
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif movimiento.stock_nuevo < movimiento.stock_anterior:
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            # Motivo
            cell = ws.cell(row=row_num, column=8, value=movimiento.motivo or '')
            cell.border = border_style
            cell.font = Font(size=9)
            cell.alignment = left_align
            
            # Usuario
            cell = ws.cell(row=row_num, column=9, value=movimiento.usuario)
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            ws.row_dimensions[row_num].height = 18
        
        # Ajustar ancho de columnas
        column_widths = [18, 30, 15, 15, 12, 15, 15, 30, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'movimientos_stock_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class PedidoProveedorViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para gestionar el historial de pedidos a proveedores"""
    queryset = PedidoProveedor.objects.all().select_related('proveedor')
    serializer_class = PedidoProveedorSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Permite filtrar por proveedor y estado"""
        queryset = PedidoProveedor.objects.all().select_related('proveedor')
        
        proveedor_id = self.request.query_params.get('proveedor', None)
        if proveedor_id:
            queryset = queryset.filter(proveedor_id=proveedor_id)
        
        estado = self.request.query_params.get('estado', None)
        if estado:
            queryset = queryset.filter(estado=estado)
        
        fecha_desde = self.request.query_params.get('fecha_desde', None)
        if fecha_desde:
            queryset = queryset.filter(fecha_envio__date__gte=fecha_desde)
        
        fecha_hasta = self.request.query_params.get('fecha_hasta', None)
        if fecha_hasta:
            queryset = queryset.filter(fecha_envio__date__lte=fecha_hasta)
        
        return queryset.order_by('-fecha_envio')

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exportar historial de pedidos a Excel con diseño mejorado"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from datetime import datetime
        
        queryset = self.get_queryset()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Pedidos"
        
        # Estilos mejorados
        title_fill = PatternFill(start_color="1a5f1a", end_color="1a5f1a", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        data_font = Font(size=10)
        number_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Título del reporte
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = 'HISTORIAL DE PEDIDOS A PROVEEDORES'
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.border = border_style
        ws.row_dimensions[1].height = 25
        
        # Información del reporte
        ws.merge_cells('A2:J2')
        info_cell = ws['A2']
        info_cell.value = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")} | Total de registros: {queryset.count()}'
        info_cell.font = Font(size=9, italic=True, color='666666')
        info_cell.alignment = center_align
        ws.row_dimensions[2].height = 18
        
        # Encabezados
        headers = [
            'ID', 'Fecha Envío', 'Proveedor', 'Email', 'Estado', 
            'Cant. Productos', 'Total Items', 'Fecha Estimada Entrega', 
            'Usuario', 'Notas'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        ws.row_dimensions[3].height = 20
        
        # Datos con formato mejorado
        for row_num, pedido in enumerate(queryset, 4):
            # ID
            cell = ws.cell(row=row_num, column=1, value=pedido.id)
            cell.border = border_style
            cell.font = number_font
            cell.alignment = center_align
            
            # Fecha Envío
            cell = ws.cell(row=row_num, column=2, value=pedido.fecha_envio.strftime('%d/%m/%Y %H:%M'))
            cell.border = border_style
            cell.font = data_font
            cell.alignment = center_align
            
            # Proveedor
            cell = ws.cell(row=row_num, column=3, value=pedido.proveedor.nombre)
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Email
            cell = ws.cell(row=row_num, column=4, value=pedido.email_enviado or '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Estado con color
            estado = pedido.get_estado_display()
            cell = ws.cell(row=row_num, column=5, value=estado)
            cell.border = border_style
            cell.font = Font(size=10, bold=True)
            cell.alignment = center_align
            # Color según estado
            if pedido.estado == 'ENVIADO':
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif pedido.estado == 'CONFIRMADO':
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif pedido.estado == 'RECIBIDO':
                cell.fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
            elif pedido.estado == 'CANCELADO':
                cell.fill = PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid")
            
            # Cantidad Productos
            cell = ws.cell(row=row_num, column=6, value=len(pedido.items) if pedido.items else 0)
            cell.border = border_style
            cell.font = number_font
            cell.alignment = center_align
            
            # Total Items
            cell = ws.cell(row=row_num, column=7, value=pedido.total_items)
            cell.border = border_style
            cell.font = number_font
            cell.alignment = center_align
            
            # Fecha Estimada
            cell = ws.cell(row=row_num, column=8, value=pedido.fecha_estimada_entrega.strftime('%d/%m/%Y') if pedido.fecha_estimada_entrega else '')
            cell.border = border_style
            cell.font = data_font
            cell.alignment = center_align
            
            # Usuario
            cell = ws.cell(row=row_num, column=9, value=pedido.usuario)
            cell.border = border_style
            cell.font = data_font
            cell.alignment = left_align
            
            # Notas
            cell = ws.cell(row=row_num, column=10, value=pedido.notas or '')
            cell.border = border_style
            cell.font = Font(size=9)
            cell.alignment = left_align
            ws.row_dimensions[row_num].height = 18
        
        # Ajustar ancho de columnas
        column_widths = [8, 18, 28, 30, 18, 15, 12, 20, 18, 35]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Crear segunda hoja con detalles de productos
        ws2 = wb.create_sheet(title="Detalle Productos")
        
        # Título segunda hoja
        ws2.merge_cells('A1:G1')
        title_cell2 = ws2['A1']
        title_cell2.value = 'DETALLE DE PRODUCTOS POR PEDIDO'
        title_cell2.fill = title_fill
        title_cell2.font = title_font
        title_cell2.alignment = center_align
        title_cell2.border = border_style
        ws2.row_dimensions[1].height = 25
        
        # Encabezados detalle
        headers_detalle = ['ID Pedido', 'Fecha', 'Proveedor', 'Código', 'Producto', 'Cantidad', 'Unidad']
        for col_num, header in enumerate(headers_detalle, 1):
            cell = ws2.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
        ws2.row_dimensions[2].height = 20
        
        # Datos detalle
        row_num = 3
        for pedido in queryset:
            if pedido.items:
                for item in pedido.items:
                    ws2.cell(row=row_num, column=1, value=pedido.id).border = border_style
                    ws2.cell(row=row_num, column=1).font = number_font
                    ws2.cell(row=row_num, column=1).alignment = center_align
                    
                    ws2.cell(row=row_num, column=2, value=pedido.fecha_envio.strftime('%d/%m/%Y %H:%M')).border = border_style
                    ws2.cell(row=row_num, column=2).font = data_font
                    ws2.cell(row=row_num, column=2).alignment = center_align
                    
                    ws2.cell(row=row_num, column=3, value=pedido.proveedor.nombre).border = border_style
                    ws2.cell(row=row_num, column=3).font = data_font
                    ws2.cell(row=row_num, column=3).alignment = left_align
                    
                    ws2.cell(row=row_num, column=4, value=item.get('codigo', '')).border = border_style
                    ws2.cell(row=row_num, column=4).font = data_font
                    ws2.cell(row=row_num, column=4).alignment = left_align
                    
                    ws2.cell(row=row_num, column=5, value=item.get('nombre', '')).border = border_style
                    ws2.cell(row=row_num, column=5).font = data_font
                    ws2.cell(row=row_num, column=5).alignment = left_align
                    
                    ws2.cell(row=row_num, column=6, value=item.get('cantidad', 0)).border = border_style
                    ws2.cell(row=row_num, column=6).font = number_font
                    ws2.cell(row=row_num, column=6).alignment = center_align
                    
                    ws2.cell(row=row_num, column=7, value=item.get('unidad_medida', '')).border = border_style
                    ws2.cell(row=row_num, column=7).font = data_font
                    ws2.cell(row=row_num, column=7).alignment = center_align
                    
                    ws2.row_dimensions[row_num].height = 18
                    row_num += 1
        
        # Ajustar ancho de columnas detalle
        column_widths_detalle = [12, 18, 28, 18, 35, 12, 12]
        for col_num, width in enumerate(column_widths_detalle, 1):
            ws2.column_dimensions[get_column_letter(col_num)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'historial_pedidos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

